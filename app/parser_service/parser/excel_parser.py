"""Deterministic Excel parsing with cell-level provenance.

Uses openpyxl (not pandas) for the parse step specifically because we need the
real A1-style cell address of every value to build citations. pandas would drop
that. Compute/aggregation (a later phase) can use pandas on top of these Values.

Assumption (PoC): synthetic files are clean — header is a single row (default
row 1). Messy real-world headers (merged cells, multi-row) are out of scope here
and tracked as a later hardening item.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .schema import CellRef, Table, Value


def parse_workbook(path, header_row: int = 1) -> dict:
    """Parse an .xlsx into provenance-tagged tables, one per worksheet.

    Returns ``{sheet_name: Table}``. ``header_row`` is 1-based.
    """
    path = Path(path)
    fname = path.name
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        return {ws.title: _parse_sheet(ws, fname, header_row) for ws in wb.worksheets}
    finally:
        wb.close()


def _parse_sheet(ws, fname: str, header_row: int) -> Table:
    all_rows = list(ws.iter_rows())
    if len(all_rows) < header_row:
        return Table(file=fname, sheet=ws.title, headers=[], rows=[])

    header_cells = all_rows[header_row - 1]
    headers = [("" if c.value is None else str(c.value)) for c in header_cells]

    data_rows: list = []
    for cells in all_rows[header_row:]:
        if all(c.value is None for c in cells):
            continue  # skip fully-empty rows deterministically
        record: dict = {}
        for idx, c in enumerate(cells):
            header = headers[idx] if idx < len(headers) and headers[idx] else None
            col_letter = get_column_letter(c.column)
            ref = CellRef(
                file=fname,
                sheet=ws.title,
                cell=c.coordinate,
                row=c.row,
                column=col_letter,
                header=header,
            )
            key = header if header else col_letter
            record[key] = Value(value=c.value, ref=ref)
        data_rows.append(record)

    return Table(file=fname, sheet=ws.title, headers=headers, rows=data_rows)
